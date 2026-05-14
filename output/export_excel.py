"""
output/export_excel.py
Writes the aggregated inventory data to a formatted, pivot-ready Excel workbook.
"""

import os
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "exports")
os.makedirs(OUTPUT_DIR, exist_ok=True)

HEADER_FILL = PatternFill("solid", start_color="1F3864")   # dark navy
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Arial", size=10)
ALT_FILL = PatternFill("solid", start_color="EEF2F7")      # light blue-grey
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_sheet(ws, df: pd.DataFrame):
    """Apply formatting to a worksheet after data has been written."""
    for col_idx, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        cell = ws[f"{col_letter}1"]
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.row_dimensions[1].height = 28

        # Auto-size columns
        max_len = max(
            len(str(col_name)),
            df[col_name].astype(str).str.len().max() if len(df) > 0 else 0,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    for row_idx in range(2, len(df) + 2):
        fill = ALT_FILL if row_idx % 2 == 0 else None
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def export(full_df: pd.DataFrame, summary_df: pd.DataFrame) -> str:
    """
    Write two sheets to an Excel file:
      - 'All Inventory'  : full normalized records (pivot-ready)
      - 'SKU Summary'    : one row per SKU with qty by source + total

    Returns the path to the saved file.
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"inventory_{timestamp}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)

    # Drop the raw column before exporting
    export_df = full_df.drop(columns=["raw"], errors="ignore")

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        export_df.to_excel(writer, sheet_name="All Inventory", index=False)
        summary_df.to_excel(writer, sheet_name="SKU Summary", index=False)

    wb = load_workbook(filepath)

    _style_sheet(wb["All Inventory"], export_df)
    _style_sheet(wb["SKU Summary"], summary_df)

    wb.save(filepath)
    return filepath
